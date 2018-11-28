#!/usr/bin/python
# -*- coding: UTF-8 -*-
'''
Created on Nov 27, 2018

@author: ashish.maikhuri
'''
import os
import openpyxl as o
os.chdir('D:\\test')
from docx  import Document
skiplist=[]
def datafill():
    wb =o.load_workbook('incident.xlsx')
    sh=wb['Page 1']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Short description'):
            des=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Number'):
            Ino=i
            break
    
    templ=['Bitlocker','Dell ITA','OKTA','Splunk','IPS - Intrusion prevention system','Dell Total Privileged Access Management',
           'OKTA Dir SYNC','Symantec','Scansafe','Dell Data Protection']
    temp2=['1:Bitlocker','2:OKTA','3:Splunk','4:IPS - Intrusion prevention system','5:Dell Total Privileged Access Management',
           '6:Symantec','7:Scansafe','8:Dell Data Protection','9:Skip the data']
    for i in range(2,sh.max_row+1):
        if not(sh.cell(i,subcolumn).value) or sh.cell(i,subcolumn).value not in templ:
            print('Sorry for inconvinence But this data is not what I expected')
            print('can you help?')
            print(sh.cell(i,Ino).value)
            print(sh.cell(i,des).value)
            print('subservices: '+sh.cell(i,subcolumn).value)
            print('please enter according to index' )
            print(temp2)
            flag=input()
            if(int(flag)!=9):
                sh.cell(i,subcolumn).value=temp2[int(flag)-1][2:]
            else:
                skiplist.append(i)       
    wb.save('new.xlsx')         
    return True
def totalincident():
    count=0
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    temp3=['IPS � Intrusion prevention system','Splunk']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value not in temp3 and i not in skiplist):
            count+=1
    return 'There are {} infra tickets in our queue.'.format(count)
            

def AUFincident():
    count=0
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    temp3=['IPS � Intrusion prevention system','Splunk']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value not in temp3 and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count+=1
    return 'Awaiting user info  : {}'.format(count)


def Assignedincident():
    count=0
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    temp3=['IPS � Intrusion prevention system','Splunk']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value not in temp3 and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count+=1
    return 'Assigned            : {}'.format(count)

def wipincident():
    count=0
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    temp3=['IPS � Intrusion prevention system','Splunk']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value not in temp3 and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count+=1
    return 'Work in progress    : {}'.format(count) 

def A3partyincident():
    count=0
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    temp3=['IPS � Intrusion prevention system','Splunk']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value not in temp3 and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count+=1
    return 'Awaiting 3rd party  : {}'.format(count)  
def Achangeincident():
    count=0
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    temp3=['IPS � Intrusion prevention system','Splunk']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value not in temp3 and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count+=1
    return 'Awaiting change     : {}'.format(count)   

def allreq():
    wb=o.load_workbook('sc_req_item.xlsx')
    sh=wb['Page 1']
    return 'There are {} requests in our queue'.format(sh.max_row-1)
def awatinguserinfotable():
    count=[0]*10
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value in ['Dell Data Protection','Dell ITA'] and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[0]+=1
        elif(sh.cell(i,subcolumn).value=='Bitlocker' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[1]+=1
        elif(sh.cell(i,subcolumn).value=='Symantec' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[2]+=1
        elif(sh.cell(i,subcolumn).value=='Splunk' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[3]+=1
        elif(sh.cell(i,subcolumn).value in ['OKTA','OKTA Dir SYNC'] and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[4]+=1
        elif(sh.cell(i,subcolumn).value=='Scansafe' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[5]+=1
        elif(sh.cell(i,subcolumn).value=='IPS � Intrusion prevention system' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[6]+=1
        elif(sh.cell(i,subcolumn).value=='Dell Total Privileged Access Management' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[7]+=1
        elif(sh.cell(i,subcolumn).value=='VR' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting User Info'):
                count[8]+=1
    sum=0
    for i in range(0,9):
        sum=sum+count[i]
    count[9]=sum
    return count
def assignedtable():
    count=[0]*10
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value in ['Dell Data Protection','Dell ITA'] and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[0]+=1
        elif(sh.cell(i,subcolumn).value=='Bitlocker' and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[1]+=1
        elif(sh.cell(i,subcolumn).value=='Symantec' and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[2]+=1
        elif(sh.cell(i,subcolumn).value=='Splunk' and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[3]+=1
        elif(sh.cell(i,subcolumn).value in ['OKTA','OKTA Dir SYNC'] and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[4]+=1
        elif(sh.cell(i,subcolumn).value=='Scansafe' and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[5]+=1
        elif(sh.cell(i,subcolumn).value=='IPS � Intrusion prevention system' and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[6]+=1
        elif(sh.cell(i,subcolumn).value=='Dell Total Privileged Access Management' and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[7]+=1
        elif(sh.cell(i,subcolumn).value=='VR' and i not in skiplist):
            if(sh.cell(i,state).value=='Assigned'):
                count[8]+=1
    sum=0
    for i in range(0,9):
        sum=sum+count[i]
    count[9]=sum
    return count

def aw3rdpartytable():
    count=[0]*10
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value in ['Dell Data Protection','Dell ITA'] and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[0]+=1
        elif(sh.cell(i,subcolumn).value=='Bitlocker' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[1]+=1
        elif(sh.cell(i,subcolumn).value=='Symantec' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[2]+=1
        elif(sh.cell(i,subcolumn).value=='Splunk' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[3]+=1
        elif(sh.cell(i,subcolumn).value in ['OKTA','OKTA Dir SYNC'] and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[4]+=1
        elif(sh.cell(i,subcolumn).value=='Scansafe' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[5]+=1
        elif(sh.cell(i,subcolumn).value=='IPS � Intrusion prevention system' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[6]+=1
        elif(sh.cell(i,subcolumn).value=='Dell Total Privileged Access Management' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[7]+=1
        elif(sh.cell(i,subcolumn).value=='VR' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting 3rd Party'):
                count[8]+=1
    sum=0
    for i in range(0,9):
        sum=sum+count[i]
    count[9]=sum
    return count
    
def wiptable():
    count=[0]*10
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value in ['Dell Data Protection','Dell ITA'] and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[0]+=1
        elif(sh.cell(i,subcolumn).value=='Bitlocker' and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[1]+=1
        elif(sh.cell(i,subcolumn).value=='Symantec' and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[2]+=1
        elif(sh.cell(i,subcolumn).value=='Splunk' and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[3]+=1
        elif(sh.cell(i,subcolumn).value in ['OKTA','OKTA Dir SYNC'] and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[4]+=1
        elif(sh.cell(i,subcolumn).value=='Scansafe' and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[5]+=1
        elif(sh.cell(i,subcolumn).value=='IPS � Intrusion prevention system' and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[6]+=1
        elif(sh.cell(i,subcolumn).value=='Dell Total Privileged Access Management' and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[7]+=1
        elif(sh.cell(i,subcolumn).value=='VR' and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[8]+=1
    sum=0
    for i in range(0,9):
        sum=sum+count[i]
    count[9]=sum
    return count
def awchangetable():
    count=[0]*10
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,subcolumn).value in ['Dell Data Protection','Dell ITA'] and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count[0]+=1
        elif(sh.cell(i,subcolumn).value=='Bitlocker' and i not in skiplist):
            if(sh.cell(i,state).value=='Work In Progress'):
                count[1]+=1
        elif(sh.cell(i,subcolumn).value=='Symantec' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count[2]+=1
        elif(sh.cell(i,subcolumn).value=='Splunk' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count[3]+=1
        elif(sh.cell(i,subcolumn).value in ['OKTA','OKTA Dir SYNC'] and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count[4]+=1
        elif(sh.cell(i,subcolumn).value=='Scansafe' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count[5]+=1
        elif(sh.cell(i,subcolumn).value=='IPS � Intrusion prevention system' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count[6]+=1
        elif(sh.cell(i,subcolumn).value=='Dell Total Privileged Access Management' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count[7]+=1
        elif(sh.cell(i,subcolumn).value=='VR' and i not in skiplist):
            if(sh.cell(i,state).value=='Awaiting change'):
                count[8]+=1
    sum=0
    for i in range(0,9):
        sum=sum+count[i]
    count[9]=sum
    return count
def reqtable():
    wb=o.load_workbook('sc_req_item.xlsx')
    sh=wb['Page 1']
    count=[0]*6
    wb=o.load_workbook('new.xlsx')
    sh=wb['Page 1']
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='Sub service'):
            subcolumn=i
            break
    for i in range(1,sh.max_column+1):
        if(sh.cell(1,i).value=='State'):
            state=i
            break
    for i in range(2,sh.max_row+1):
        if(sh.cell(i,state).value=='Open'):
            count[0]+=1
        if(sh.cell(i,state).value=='Awaiting 3rd Party'):
            count[1]+=1
        if(sh.cell(i,state).value=='Awaiting User Info'):
            count[2]+=1
        if(sh.cell(i,state).value=='Work In Progress'):
            count[3]+=1
        if(sh.cell(i,state).value=='Assigned'):
            count[4]+=1
    sum=0
    for i in range(0,5):
        sum=sum+count[i]
    count[5]=sum
    return count
    
